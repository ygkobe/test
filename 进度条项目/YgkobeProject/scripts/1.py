import sys
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


class OpenChromeBaidu(object):

    def __init__(self):
        # 设置要访问的百度翻译页面
        self.url = "https://fanyi.baidu.com/?fr=pcPinzhuan#en/zh/"

        # 初始化 Chrome 浏览器的选项
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--lang=en-US")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
        )

        # 创建 Chrome WebDriver 对象并传入选项
        self.driver = webdriver.Chrome(options=chrome_options)

        # 创建 Excel 工作簿
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Translations"
        self.sheet.append(["Search Text", "Translation", "Screenshot"])

        # 设置列宽和列高度，可以根据实际需求调整数值
        self.sheet.column_dimensions['A'].width = 30
        # 使用整数索引设置列高度
        # self.sheet.row_dimensions[1].height = 100
        self.sheet.row_dimensions[2].height = 260
        self.sheet.row_dimensions[3].height = 260
        # # 设置第一列的所有行高度为 100，可根据需求调整
        # for i in range(1, self.sheet.max_row + 1):
        #     self.sheet.row_dimensions[i].height = 100

        self.sheet.column_dimensions['B'].width = 50

        self.sheet.column_dimensions['C'].width = 90

    def take_screenshot_and_save(self, row_num):
        # 截图保存为 PNG 文件
        screenshot_path = f"screenshot_{row_num}.png"
        self.driver.save_screenshot(screenshot_path)

        # 将图片调整为合适的大小并插入到 Excel 中
        img = PILImage.open(screenshot_path)
        img = img.resize((300, 250))
        img.save(screenshot_path)

        img = Image(screenshot_path)
        self.sheet.add_image(img, f'C{row_num}')

    def run(self):
        # 打开百度翻译页面
        self.driver.get(self.url)
        # 打开页面优先执行的 js,execute_cdp_cmd
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                    Object.defineProperty(navigator, 'webdriver', {
                      get: () => undefined
                    })
                  """
        })
        # 翻译第一个文本
        select_input = WebDriverWait(self.driver, 5, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="editor-text"]/div[1]/div[1]/div/div/div/div'))
        )
        search_text = '巴黎世家'
        select_input.send_keys(search_text + Keys.RETURN)
        time.sleep(1)

        result_value = WebDriverWait(self.driver, 5, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="trans-selection"]/div/span'))
        )
        translation_result = result_value.text
        print(translation_result)

        # 将结果保存到 Excel
        self.sheet.append([search_text, translation_result])
        self.take_screenshot_and_save(self.sheet.max_row)

        select_input.clear()
        time.sleep(2)

        # 翻译第二个文本
        search_text = '我还能说些什么呢'
        select_input = WebDriverWait(self.driver, 5, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="editor-text"]/div[1]/div[1]/div/div/div/div'))
        )
        select_input.send_keys(search_text + Keys.RETURN)
        time.sleep(1)

        result_value = WebDriverWait(self.driver, 5, 0.5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="trans-selection"]/div/span'))
        )
        translation_result = result_value.text
        print(translation_result)

        # 将结果保存到 Excel
        self.sheet.append([search_text, translation_result])
        self.take_screenshot_and_save(self.sheet.max_row)

        # 保存 Excel 文件
        self.workbook.save("Baidu_Translations.xlsx")


def main():
    # 初始化并运行自动化操作
    open_browser = OpenChromeBaidu()
    open_browser.run()


if __name__ == '__main__':
    main()