import io
import time
from celery import shared_task
from openpyxl import Workbook
import random
# -*- coding:utf8 -*-
import time
import json

from django.db import transaction
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import StreamingHttpResponse
from rest_framework.permissions import AllowAny  # 导入需要的权限类
from rest_framework import viewsets
from rest_framework.decorators import action
from django.conf import settings
import datetime

import pandas as pd
import pymysql
from datetime import datetime

from django.http import HttpResponse
import requests
from io import BytesIO
from openpyxl import Workbook
from collections import defaultdict
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font
import warnings
# 抑制特定的警告
warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy connectable.*")


# 获取省市字段
def extract_city(address: dict):
    """
    获取省市
    :param address:
    :return:
    """
    offer = address.get('offer')
    offer_address = address.get('address')
    if offer and offer_address:

        if "省" in offer:
            return offer.split("省")[0] + "省"
        if "省" in offer_address:
            return offer_address.split("省")[0] + "省"

        if "市" in offer:
            return offer.split("市")[0] + "市"

        if "市" in offer_address:
            return offer_address.split("市")[0] + "市"

        if "县" in offer:
            return offer.split("县")[0] + "县"
        elif "县" in offer_address:
            return offer_address.split("县")[0] + "县"
        else:
            return "未知地点"


# 查询数据库 获取汽车列表
def get_database_car_numbers(group_name=None, create_time=None):
    """
    查询数据库车辆数据
    :param group_name:
    :param create_time:
    :return:
    """
    config = settings.DATABASES.get('default')
    host = config.get('HOST')
    port = config.get('PORT')
    password = config.get('PASSWORD')
    database = config.get('NAME')
    user = config.get('USER')

    sql = """
    select car_number 
    from car_numbers_group 
    WHERE 
    group_name = '{}' AND create_time = '{}'
    """.format(group_name, create_time)

    # 建立数据库连接
    conn = pymysql.connect(host=host, user=user, password=password, database=database, port=port)

    # 编写 SQL 查询语句

    # 使用 pandas 的 read_sql 函数读取数据
    data = pd.read_sql(sql, conn)
    data = data.fillna('').to_dict('records')

    car_list = list(set([number.get('car_number') for number in data]))

    return car_list


def traffic_violation_view(car_number=None, startDate=None, endDate=None, cookies=None, Host=None, Origin=None, Referer=None, url=None):
    """
    交通数据概览
    :param car_number:
    :return:
    """
    payload = {
        "startDate": startDate,
        "endDate": endDate,
        "hphm": car_number,
        "hpzl": 52,
    }

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
        "Cookie": cookies,
        "Host": Host,
        "Origin": Origin,
        "Referer": Referer,
        "Sec-Ch-Ua": '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
        "Sec-Ch-Mobile": '?0',
        "Sec-Ch-Ua-Platform": "macOS"
    }
    response = requests.post(url=url, data=payload, headers=headers, timeout=3)

    data = json.loads(response.text)
    # print(data)

    if data.get('code') == 200:

        if data.get('data').get('content'):
            data_list = []
            # print(data)
            for item in data.get('data').get('content'):
                # print(i)
                jkbj = item.get("jkbj")
                clbj = item.get("clbj")
                # data_list.append(item)
                if int(jkbj) == 0 and int(clbj) == 0:
                    data_list.append(item)

            if len(data_list) >= 1:
                return data_list
            else:
                return [{'hphm': car_number, "info": ""}]

        else:
            return [{'hphm': car_number, "info": ""}]

    elif data.get('code') == 500 and data.get('message') == "只能查询已绑定的机动车违法":
        return [{'hphm': car_number, 'info': ': 未绑定'}]
    else:
        return [{'hphm': car_number, 'info': ': 未知错误'}]


def traffic_violation_details(car_number=None, xh=None, cjjg=None, cookies=None, Host=None, Origin=None, Referer=None, url=None):
    """
    交通数据详细信息
    :param car_number:
    :param xh:
    :param cjjg:
    :return:
    """
    payload = {
        "hphm": car_number,
        "hpzl": 52,
        "xh": xh,
        "cjjg": cjjg

    }

    headers = {
        "Accept": "application/json, text/javascript, */*; q=0.01",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
        "Cookie": cookies,
        "Host": Host,
        "Origin": Origin,
        "Referer": Referer,
        "Sec-Ch-Ua": '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
        "Sec-Ch-Mobile": '?0',
        "Sec-Ch-Ua-Platform": "macOS",
        "Connection": "keep-alive"
    }
    # response = requests.post(url=url, data=payload, headers=headers)
    response = requests.post(url=url, data=payload, headers=headers,timeout=3)
    # print(response.text)

    response_status_code = json.loads(response.text).get('code')
    if response_status_code == 200:
        data = json.loads(response.text).get('data')
        return data


def get_grouped_data_list(car_number_list=None, cookies=None, startDate=None, endDate=None):
    """
    获取分组数据
    :param self:
    :return:
    """
    print('获取分组数据')
    data_list = []

    # for car_number in get_database_car_numbers(group_name=group_name, create_time=create_time):
    for car_number in car_number_list:
        print(car_number)
        if car_number.startswith('粤'):
            Host = "gd.122.gov.cn"
            Origin = "https://gd.122.gov.cn"
            Referer = "https://gd.122.gov.cn/views/memfyy/violation.html"
            url = "https://gd.122.gov.cn/user/m/uservio/suriquery"
            details_url = "https://gd.122.gov.cn/user/m/tsc/vio/querySurvielDetail"
        elif car_number.startswith('京'):
            Host = "bj.122.gov.cn"
            Origin = "https://bj.122.gov.cn"
            Referer = "https://bj.122.gov.cn/views/memfyy/violation.html"
            url = "https://bj.122.gov.cn/user/m/uservio/suriquery"
            details_url = "https://bj.122.gov.cn/user/m/tsc/vio/querySurvielDetail"
        else:
            Host = "gd.122.gov.cn"
            Origin = "https://gd.122.gov.cn"
            Referer = "https://gd.122.gov.cn/views/memfyy/violation.html"
            url = "https://gd.122.gov.cn/user/m/uservio/suriquery"
            details_url = "https://gd.122.gov.cn/user/m/tsc/vio/querySurvielDetail"
        # print(car_number)

        try:
            # 概览数据
            view_data = traffic_violation_view(
                car_number=car_number, startDate=startDate, endDate=endDate,
                cookies=cookies, Host=Host, Origin=Origin, Referer=Referer, url=url
            )

            for detail_data in view_data:
                print(detail_data)
                xh = detail_data.get('xh')
                cjjg = detail_data.get('cjjg')
                if xh and cjjg:
                    data = traffic_violation_details(car_number=car_number, xh=xh, cjjg=cjjg, cookies=cookies, Host=Host, Origin=Origin, Referer=Referer, url=details_url)
                    print(data)
                    data_list.append(data)
                else:

                    data = {"hphm": car_number + detail_data.get('info')}
                    print(data)
                    data_list.append(data)
        except Exception as e:
            data_list.append({'hphm': car_number + ": 数据错误"})

    grouped_data = defaultdict(list)

    for entry in data_list:

        if entry.get('wjgbj'):
            wjgbj = entry.get('wjgbj')
        else:
            if entry.get('wfms'):
                wjgbj = 0
            else:
                wjgbj = ''

        item = {
            'wfms': entry.get('wfms'),
            'wfsj': entry.get('wfsj'),
            'wfdz': entry.get('wfdz'),
            'wjgbj': wjgbj,
            'fkje': entry.get('fkje'),
            'cjjgmc': entry.get('cjjgmc'),
            'hpzlStr': entry.get('hpzlStr'),
            'photos': entry.get('photos')[0] if entry.get('photos') else None,
        }
        grouped_data[entry['hphm']].append(item)
    return grouped_data


def generate_excel(grouped_data):
    grouped_data = grouped_data
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '交通违法数据'
    sheet.column_dimensions['A'].width = 17.0
    sheet.column_dimensions['B'].width = 17.0
    sheet.column_dimensions['C'].width = 10.0
    sheet.column_dimensions['D'].width = 40.0
    sheet.column_dimensions['E'].width = 20.0
    sheet.column_dimensions['F'].width = 35.0
    sheet.column_dimensions['G'].width = 10.0
    sheet.column_dimensions['H'].width = 10.0
    sheet.column_dimensions['I'].width = 25.0
    sheet.column_dimensions['J'].width = 41.0
    sheet.column_dimensions['K'].width = 36.0

    # title header
    header_fill = PatternFill(start_color="43a6f2", end_color="43a6f2", fill_type="solid")

    header_cells = ["序号", "车牌号", "编号", "违章时间", "违章地点", "违章名称", "罚分", "罚款", "违法城市","提供方", "违章照片"]

    for col_num, header in enumerate(header_cells, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.fill = header_fill

    # 调整第一行高度
    sheet.row_dimensions[1].height = 35

    # 写入到excel
    serial_id = 1  # 序号id
    row = 2  # 内容起始行
    for hphm, wfsj_list in grouped_data.items():
        start_row = row
        end_row = row + len(wfsj_list) - 1
        private_id = 1  # 编号id
        for item in wfsj_list:
            sheet.cell(row=row, column=1, value=serial_id)
            sheet.cell(row=row, column=2, value=hphm).alignment = Alignment(wrap_text=True,
                                                                                 horizontal='center',
                                                                                 vertical='center')

            if item.get('wfsj') and item.get('wfms'):

                sheet.cell(row=row, column=3, value=private_id).alignment = Alignment(wrap_text=True,
                                                                                           horizontal='center',
                                                                                           vertical='center')
            else:
                sheet.cell(row=row, column=3, value="无违法").alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')

            sheet.cell(row=row, column=4, value=item['wfsj']).alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')
            sheet.cell(row=row, column=5, value=item['wfdz']).alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')
            sheet.cell(row=row, column=6, value=item['wfms']).alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')
            sheet.cell(row=row, column=7, value=item.get('wjgbj')).alignment = Alignment(wrap_text=True,
                                                                                              horizontal='center',
                                                                                              vertical='center')
            sheet.cell(row=row, column=8, value=item['fkje']).alignment = Alignment(wrap_text=True,
                                                                                         horizontal='center',
                                                                                         vertical='center')
            wf_city = extract_city({'offer': item['cjjgmc'], 'address': item['wfdz']})
            sheet.cell(row=row, column=9, value=wf_city).alignment = Alignment(wrap_text=True,
                                                                                    horizontal='center',
                                                                                    vertical='center')
            sheet.cell(row=row, column=10, value=item['cjjgmc']).alignment = Alignment(wrap_text=True,
                                                                                            horizontal='center',
                                                                                            vertical='center')
            try:
                if item['photos']:

                    img_url = item['photos']
                    img_data = None
                    retries = 3
                    for attempt in range(retries):
                        try:

                            headers = {
                                "Accept": "application/json, text/javascript, */*; q=0.01",
                                "Accept-Encoding": "gzip, deflate, br, zstd",
                                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                                'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36",
                                "Sec-Ch-Ua": '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
                                "Sec-Ch-Mobile": '?0',
                                "Sec-Ch-Ua-Platform": "macOS"
                            }
                            response = requests.get(img_url, timeout=3, headers=headers)
                            response.raise_for_status()
                            img_data = response.content
                            break
                        except requests.RequestException:
                            if attempt < retries - 1:
                                continue
                            else:

                                # 插入图片地址
                                sheet.cell(row=row, column=11, value=img_url)

                                # 获取单元格对象
                                cell = sheet.cell(row=row, column=11)

                                # 插入超链接
                                cell.hyperlink = img_url

                                # 应用对齐方式
                                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

                                # 设置字体颜色
                                font = Font(color="3FA9FF")
                                cell.font = font

                    if img_data:
                        img = Image(BytesIO(img_data))
                        # 调整图片大小以适应单元格
                        img.width = 210  # 宽度
                        img.height = 140  # 高度
                        img_cell = sheet.cell(row=row, column=11)
                        sheet.add_image(img, img_cell.coordinate)
                else:

                    sheet.cell(row=row, column=11, value=item.get('photos'))
            except Exception as e:

                sheet.cell(row=row, column=11, value="{}-{}".format(e, item.get('photos')))
            # 内部id
            private_id = private_id + 1
            row += 1

        sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
        top_left_cell = sheet.cell(row=start_row, column=1)
        top_left_cell.value = serial_id
        top_left_cell.alignment = Alignment(horizontal='center', vertical='center')

        # sheet.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
        # merged_cell = sheet.cell(row=start_row, column=2)
        # merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Move to the next available row
        row = end_row + 1
        serial_id = serial_id + 1

    # 设置第二行开始的行高为 3.9cm
    for i in range(2, row):
        sheet.row_dimensions[i].height = 3.9 * 28.3465  # 将厘米转换为点

    # 将工作簿保存到内存中
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


@shared_task(bind=True)
def long_running_task(self, form_data):
    # 模拟长时间任务
    print("开始任务", form_data)
    # {'group_id': 1, 'date_range': ['2024-05-19', '2024-08-17'], 'token': '3wqeqw', 'group_name': '哈喽','create_time': '2024-08-12 22:15:03'}

    start_date = form_data.get('date_range')[0].replace('-', '')
    end_date = form_data.get('date_range')[1].replace('-', '')
    cookies = form_data.get('token')
    group_name = form_data.get('group_name')
    create_time = form_data.get('create_time')

    print("处理车牌号列表")
    self.update_state(state='PROGRESS', meta={'current': 20})
    car_numbers = get_database_car_numbers(group_name=group_name, create_time=create_time)
    print(car_numbers)
    time.sleep(1)

    self.update_state(state='PROGRESS', meta={'current': 40})
    time.sleep(1)

    grouped_data = get_grouped_data_list(car_number_list=car_numbers, cookies=cookies, startDate=start_date, endDate=end_date)
    self.update_state(state='PROGRESS', meta={'current': 70})
    time.sleep(1)

    file_data = generate_excel(grouped_data)
    self.update_state(state='PROGRESS', meta={'current': 80})
    time.sleep(1)

    # 将文件内容返回为二进制数据
    return {'current': 100, 'file_data': file_data}
