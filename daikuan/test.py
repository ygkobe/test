# -*- coding:utf8 -*-
import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


class FundFlowGenerate(object):

    def __init__(self):
        # 实例化 workbook
        self.work_book = Workbook()
        # 激活 worksheet
        self.work_sheet = self.work_book.active

    # 按照逻辑写入 Excel
    def generate_excel(self):

        # 设置列宽
        self.work_sheet.column_dimensions['A'].width = 20.0
        self.work_sheet.column_dimensions['B'].width = 20.0
        self.work_sheet.column_dimensions['C'].width = 9.0
        self.work_sheet.column_dimensions['D'].width = 17.0
        self.work_sheet.column_dimensions['E'].width = 25.0
        self.work_sheet.column_dimensions['F'].width = 47.0
        self.work_sheet.column_dimensions['G'].width = 25.0
        self.work_sheet.column_dimensions['H'].width = 25.0

        # 设置表格居中
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.work_sheet.title = "资金流水"

        # 合并单元格
        self.work_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        # 设置标题字体
        font = Font(color=colors.BLACK, size=14, bold=True)
        self.work_sheet.cell(row=1, column=1, value='个人客户账户明细清单').font = font
        self.work_sheet.cell(row=1, column=1).alignment = align

        # 设置行高
        self.work_sheet.row_dimensions[1].height = 50
        self.work_sheet.row_dimensions[2].height = 31
        self.work_sheet.row_dimensions[3].height = 31
        self.work_sheet.row_dimensions[4].height = 31
        self.work_sheet.row_dimensions[5].height = 26

        # 设置信息字体和对齐方式
        cell_font = Font(color=colors.BLACK, size=11)
        cell_align = Alignment(horizontal='left', vertical='center')

        # 填写客户信息
        self.work_sheet.cell(row=2, column=1, value="客户姓名 :  汪海燕").font = cell_font
        self.work_sheet.cell(row=2, column=1).alignment = cell_align
        self.work_sheet.cell(row=3, column=1, value="卡/ 账号 :  6235822000022159941").font = cell_font
        self.work_sheet.cell(row=3, column=1).alignment = cell_align
        self.work_sheet.cell(row=4, column=1, value="储       种 :  普通活期").font = cell_font
        self.work_sheet.cell(row=4, column=1).alignment = cell_align

        self.work_sheet.cell(row=3, column=3, value="分期号：001").font = cell_font
        self.work_sheet.cell(row=3, column=3).alignment = cell_align
        self.work_sheet.cell(row=4, column=3, value="存期： ").font = cell_font
        self.work_sheet.cell(row=4, column=3).alignment = cell_align

        self.work_sheet.cell(row=2, column=5, value="日期范围：2024-10-01").font = cell_font
        self.work_sheet.cell(row=2, column=5).alignment = cell_align
        self.work_sheet.cell(row=3, column=5, value="币种：人民币").font = cell_font
        self.work_sheet.cell(row=3, column=5).alignment = cell_align
        self.work_sheet.cell(row=4, column=5, value="起利日：2024-10-01").font = cell_font
        self.work_sheet.cell(row=4, column=5).alignment = cell_align

        self.work_sheet.cell(row=2, column=7, value="明细类型： 全部").font = cell_font
        self.work_sheet.cell(row=2, column=7).alignment = cell_align
        self.work_sheet.cell(row=3, column=7, value="打印日期：2024-10-01").font = cell_font
        self.work_sheet.cell(row=3, column=7).alignment = cell_align
        self.work_sheet.cell(row=4, column=7, value="钞汇标志： 钞").font = cell_font
        self.work_sheet.cell(row=4, column=7).alignment = cell_align

        # 设置表头
        body_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        body_font = Font(color=colors.BLACK, size=12, bold=True)
        headers = ['交易日期', '业务摘要', '收/支', '发生额', '余额', '对方户名', '对方账号', '交易渠道']
        for col_num, header in enumerate(headers, 1):
            cell = self.work_sheet.cell(row=5, column=col_num, value=header)
            cell.font = body_font
            cell.alignment = body_align

        # 定义边框样式，使用细边框
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 初始化余额和交易日期
        start_balance = 20020.0  # 初始金额
        init_balance = start_balance  # 初始金额
        target_balance = 500000.0  # 目标金额
        target_row = 1000
        trans_type_temp = True
        current_balance = start_balance
        transaction_date = datetime.strptime('2024-10-02', '%Y-%m-%d')

        # 交易类型和摘要   spend_range 定义了每个款项 可以消费金额的额度
        transaction_types_descriptions = [
            {
                'description': '协议支付',
                'trans_type': '支出',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': 'POS消费',
                'trans_type': '支出',
                'spend_range': range(100, 4000),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': '货款',
                'trans_type': '收入',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': '',
                'trans_type': '支出',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': '转账存入',
                'trans_type': '收入',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': '转账支出',
                'trans_type': '支出',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
            {
                'description': '日常支出',
                'trans_type': '支出',
                'spend_range': range(100, 400),
                'account_list': [
                    {
                        "account_id": "6228482122190854750",
                        "account_name": "李宁刚",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482128094466362",
                        "account_name": "姜海军",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010136576",
                        "account_name": "刘文强",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217855000082341827",
                        "account_name": "张仁娟",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217210411010134319",
                        "account_name": "黄跃书",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200006400352",
                        "account_name": "李学平",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482129149012572",
                        "account_name": "王炳富",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222621392000256392",
                        "account_name": "赵长清",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6217000200005120576",
                        "account_name": "王振廷",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6228482126263634625",
                        "account_name": "汪金旺",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6222080411000669820",
                        "account_name": "魏宏楠",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411010376372",
                        "account_name": "代洪芝",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6215580411001840633",
                        "account_name": "潘春明",
                        "transaction_channel": "支付平台渠道"
                    },
                    {
                        "account_id": "6212260411007462759",
                        "account_name": "张怀深",
                        "transaction_channel": "支付平台渠道"
                    }
                ]
            },
        ]

        row_num = 6  # 从第6行开始写入数据

        # 循环生成交易记录，直到余额达到目标值
        while current_balance < target_balance:
            # 减少交易日期，确保日期从大到小
            # days_offset = random.randint(1, 3)
            # transaction_date -= timedelta(days=days_offset)
            transaction_date -= timedelta(days=1)
            # 每日消费笔数
            daily_consumption_counts = random.randint(2, 4)
            # 定义每天消费笔数 随机 2到8笔
            for i in range(daily_consumption_counts):

                trans_type_desc = random.choice(transaction_types_descriptions)
                trans_type = trans_type_desc.get('trans_type')
                trans_type_temp = trans_type
                description = trans_type_desc.get('description')
                spend_range = trans_type_desc.get('spend_range')
                account_list = trans_type_desc.get('account_list')
                account_info = random.choice(account_list)

                account_name = account_info.get('account_name')
                account_id = account_info.get('account_id')
                transaction_channel = account_info.get('transaction_channel')

                amount = random.choice(spend_range)
                float_e = [
                    "00", "10", "20", "30", "40", "50", "60", "70", "80", "90",
                    "00", "00", "00", "00", "00", "00", "00", "00", "00", "00",
                ]
                # 优化后的代码
                # 拼接小数部分
                decimal_part = float(random.choice(float_e)) / 100
                amount = int(amount) + decimal_part
                # 保留两位小数
                amount = round(amount, 2)

                # 写入交易数据
                self.work_sheet.cell(row=row_num, column=1,
                                     value=transaction_date.strftime('%Y%m%d')).border = thin_border
                self.work_sheet.cell(row=row_num, column=2, value=description).border = thin_border
                self.work_sheet.cell(row=row_num, column=3, value=trans_type).border = thin_border
                amount_cell = self.work_sheet.cell(row=row_num, column=4, value=amount)
                amount_cell.border = thin_border
                # 设置金额列的数字格式为保留两位小数
                amount_cell.number_format = '0.00'
                # self.work_sheet.cell(row=row_num, column=4, value=amount).border = thin_border
                self.work_sheet.cell(row=row_num, column=5, value=round(current_balance, 2)).border = thin_border
                self.work_sheet.cell(row=row_num, column=6,
                                     value=account_name).border = thin_border
                self.work_sheet.cell(row=row_num, column=7, value=account_id).border = thin_border
                self.work_sheet.cell(row=row_num, column=8, value=transaction_channel).border = thin_border

                row_num += 1
                if trans_type_temp == '支出':
                    current_balance += amount
                else:
                    if current_balance - amount < 0:
                        continue  # 跳过余额为负的情况
                    if current_balance < init_balance:
                        continue
                    current_balance -= amount

        # 添加印章图片
        try:
            img = Image('印章.png')  # 请将 '印章.png' 替换为实际的图片路径
            # 设置图片大小
            img.width = 220
            img.height = 170
            self.work_sheet.add_image(img, 'H2')
        except Exception as e:
            print(e, "无图片添加")

    def save_excel(self):
        self.generate_excel()
        self.work_book.save("流水明细.xlsx")


if __name__ == '__main__':
    c = FundFlowGenerate()
    c.save_excel()
